"""离线评估 pipeline（多组版）—— ingest(物化) + judge(LLM 判定, 可选) + scoring +
aggregate，含进度事件。

路由层负责校验 metric 名与组装 groups；本模块只编排：解析记录 → 逐组逐条物化 →
（含 llm_judge 组时并发调 LLM 把 pred 分类到声明标签）→ 逐条评分 → 按组聚合，并经
``job.push_event`` 推送进度，最终把 ``EvalSummary`` 写入 ``job.result``。异常 →
``FAILED`` + error 事件（仿 ``optimize.py`` 的 ``_run_with_progress``），取消重抛。
"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import re
import traceback
from collections.abc import Awaitable, Callable
from dataclasses import dataclass, replace
from typing import Any

from openjiuwen.core.foundation.llm import Model

from evo_agent.api.jobs import Job, JobStatus
from evo_agent.evaluator.offline.judge import _judge
from evo_agent.evaluator.offline.models import (
    CaseScore,
    EvalSummary,
    GroupConfig,
    MaterializedCase,
)
from evo_agent.evaluator.offline.scorer import OfflineMetricScorer

logger = logging.getLogger(__name__)

__all__ = [
    "GroupConfig",
    "OfflineEvalRequest",
    "load_raw_records",
    "ingest",
    "run_offline_eval",
    "summary_to_dict",
]

# 进度推送粒度——每处理多少条推一次事件，避免事件风暴。
_PROGRESS_EVERY = 10

ProgressCallback = Callable[[int, int], Awaitable[None] | None]


@dataclass(frozen=True)
class OfflineEvalRequest:
    """路由层组装的离线评估请求（组配置已校验）。"""

    dataset_id: str
    raw_records: list[dict[str, Any]]
    groups: list[GroupConfig]
    id_field: str = ""
    model: Model | None = None  # 仅含 llm_judge 组时由路由层构建并注入


# ---------------------------------------------------------------------------
# 记录解析
# ---------------------------------------------------------------------------


def load_raw_records(data: bytes, filename: str | None) -> list[dict[str, Any]]:
    """按扩展名/内容探测，把上传内容解析为原始 dict 列表。

    支持 JSON 数组 / 单对象、JSONL、CSV、XLSX。返回原始 dict 列表（**不复用
    ``_load_cases``**——它硬编码 ``expected_behavior→expected_result``，与用户
    自定义字段映射冲突）。
    """
    name = (filename or "").lower()

    if name.endswith(".xlsx"):
        return _require_records(_parse_xlsx(data), "XLSX")

    text = data.decode("utf-8", errors="replace")

    if name.endswith(".csv"):
        return _parse_csv(text)
    if name.endswith(".jsonl"):
        return _require_records(_parse_jsonl(text), "JSONL")

    # .json 或未知扩展名：先试整段 JSON，再试 JSONL，最后 CSV。
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        parsed = None
    if isinstance(parsed, list):
        return _require_records(
            [r for r in parsed if isinstance(r, dict)],
            "JSON array",
        )
    if isinstance(parsed, dict):
        return [parsed]

    records = _parse_jsonl(text)
    if records:
        return records
    return _parse_csv(text)


def _require_records(records: list[dict[str, Any]], kind: str) -> list[dict[str, Any]]:
    if not records:
        raise ValueError(f"Dataset parsed as {kind} but yielded no records")
    return records


def _parse_jsonl(text: str) -> list[dict[str, Any]]:
    """每行一个 JSON 对象；跳过空行；非对象行抛错。"""
    records: list[dict[str, Any]] = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        obj = json.loads(line)  # 非法 JSON 抛 JSONDecodeError
        if not isinstance(obj, dict):
            raise ValueError("JSONL line is not a JSON object")
        records.append(obj)
    return records


def _parse_csv(text: str) -> list[dict[str, Any]]:
    """CSV → dict 列表（``csv.DictReader``）。"""
    reader = csv.DictReader(io.StringIO(text))
    records = [dict(row) for row in reader]
    if not records:
        raise ValueError("CSV yielded no records (empty or missing header)")
    return records


def _parse_xlsx(data: bytes) -> list[dict[str, Any]]:
    """XLSX → dict 列表：首个 sheet，首行为表头，其余行→dict。

    表头为 ``None`` 的列丢弃（trailing 空列）。延迟 import openpyxl 以隔离依赖。
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if all(c is None for c in row):
            continue  # 跳过全空行
        rec: dict[str, Any] = {}
        for key, value in zip(header, row):
            if key:
                rec[key] = value
        records.append(rec)
    return records


# ---------------------------------------------------------------------------
# JSON pred 解析
# ---------------------------------------------------------------------------


def _parse_json_cell(value: Any) -> dict[str, Any]:
    r"""把 pred cell 解析为 JSON 对象（dict）。

    cell 可能是 JSON 字符串、已是 dict、或 None。解析顺序（从严到松，逐级兜底）：

    1. 直接 ``json.loads``——纯 JSON cell 的快路径。
    2. `` ```json ... ``` `` 围栏正则——LLM 常把 JSON 塞在 markdown 围栏里，正则精确
       提取围栏内内容，不受散文里杂散 ``{``/``}`` 干扰。
    3. 首个 ``{`` 到末个 ``}`` 子串——无围栏、JSON 嵌在散文里的最后兜底。

    全部失败 / 非 dict / None → ValueError（fail-fast：``json_key`` 设了但该列不含
    JSON 对象）。
    """
    if isinstance(value, dict):
        return value
    if value is None:
        raise ValueError("JSON pred cell is empty")
    text = str(value)

    obj: Any = None
    try:
        obj = json.loads(text)  # 1. 纯 JSON
    except json.JSONDecodeError:
        obj = None
    if not isinstance(obj, dict):
        m = _JSON_FENCE_RE.search(text)  # 2. ```json``` 围栏
        if m:
            try:
                obj = json.loads(m.group(1))
            except json.JSONDecodeError:
                obj = None
    if not isinstance(obj, dict):
        i, j = text.find("{"), text.rfind("}")  # 3. {..} 兜底
        if i >= 0 and j > i:
            try:
                obj = json.loads(text[i : j + 1])
            except json.JSONDecodeError:
                obj = None
    if not isinstance(obj, dict):
        raise ValueError(f"pred cell is not valid JSON: {text[:80]!r}")
    return obj


# ```json ... ``` / ``` ... ``` 围栏正则（DOTALL 跨行，IGNORECASE 容 json/JSON）。
_JSON_FENCE_RE = re.compile(r"```(?:json)?\s*(.*?)\s*```", re.DOTALL | re.IGNORECASE)


def _resolve_key_path(obj: Any, key_path: str) -> Any:
    """按 ``a.b`` 点号路径从 dict 取值；任一段不是 dict 或缺键 → None。"""
    cur: Any = obj
    for part in key_path.split("."):
        if not isinstance(cur, dict) or part not in cur:
            return None
        cur = cur[part]
    return cur


# ---------------------------------------------------------------------------
# ingest（多组）
# ---------------------------------------------------------------------------


async def ingest(
    records: list[dict[str, Any]],
    groups: list[GroupConfig],
    id_field: str,
    on_progress: ProgressCallback | None = None,
) -> list[MaterializedCase]:
    """对每条 record × 每组物化。

    exact_match / llm_judge 组：``gold=record[gold_field]``（缺失 fail-fast）、
    ``extracted=record.get(pred_field)``（``json_key`` 为空）或从该 cell 的 JSON 按
    ``json_key`` 路径解析（非空）、method=``"raw"``/``"json"``。llm_judge 的
    ``judged_label`` 不在此阶段产出（见 ``_judge``），ingest 只物化 pred/gold 文本；
    llm_judge 组还需 ``labels`` 非空与 ``extract_key`` 非空（需从 pred 中提取的内容，供 judge 阶段
    渲染 prompt），缺失 fail-fast。
    keyword 组：``gold=list(keywords)``、``extracted=[命中的关键词]``（pred 文本来自
    列原值或 JSON 解析值）、method=``"keyword"``/``"json_keyword"``。
    ``case_id`` 取 ``id_field``，缺省用原始索引。
    """
    if not records:
        raise ValueError("Cannot ingest empty dataset")
    if not groups:
        raise ValueError("Cannot ingest without groups")

    first = records[0]
    for g in groups:
        if g.kind in ("exact_match", "llm_judge") and not g.gold_field:
            raise ValueError(f"Group {g.name!r} ({g.kind}) requires gold_field")
        if g.kind in ("exact_match", "llm_judge") and g.gold_field not in first:
            raise ValueError(f"gold_field {g.gold_field!r} (group {g.name!r}) not found in dataset")
        if g.pred_field not in first:
            raise ValueError(f"pred_field {g.pred_field!r} (group {g.name!r}) not found in dataset")
        if g.kind == "keyword" and not g.keywords:
            raise ValueError(f"Group {g.name!r} (keyword) requires non-empty keywords")
        if g.kind == "llm_judge" and not g.labels:
            raise ValueError(f"Group {g.name!r} (llm_judge) requires non-empty labels")
        if g.kind == "llm_judge" and not g.extract_key:
            raise ValueError(f"Group {g.name!r} (llm_judge) requires extract_key")
        if g.json_key:
            # fail-fast：json_key 设了，首记录的 pred cell 必须是 JSON 对象
            _parse_json_cell(first.get(g.pred_field))

    total = len(records) * len(groups)
    materialized: list[MaterializedCase] = []
    done = 0
    for index, record in enumerate(records):
        case_id = (
            str(record[id_field]) if id_field and record.get(id_field) is not None else str(index)
        )
        for g in groups:
            if g.json_key:
                pred_value: Any = _resolve_key_path(
                    _parse_json_cell(record.get(g.pred_field)), g.json_key
                )
            else:
                pred_value = record.get(g.pred_field)
            if g.kind == "keyword":
                gold: Any = list(g.keywords)
                pred_text = "" if pred_value is None else str(pred_value)
                extracted: Any = [k for k in g.keywords if k and k in pred_text]
                method = "json_keyword" if g.json_key else "keyword"
            else:
                gold = record.get(g.gold_field)
                extracted = pred_value
                method = "json" if g.json_key else "raw"
            materialized.append(
                MaterializedCase(
                    case_id=case_id,
                    group=g.name,
                    gold=gold,
                    extracted=extracted,
                    extraction_method=method,
                )
            )
            done += 1
            if on_progress is not None and done % _PROGRESS_EVERY == 0:
                ret = on_progress(done, total)
                if asyncio.iscoroutine(ret):
                    await ret
    if on_progress is not None:
        ret = on_progress(total, total)
        if asyncio.iscoroutine(ret):
            await ret
    return materialized


# ---------------------------------------------------------------------------
# 编排
# ---------------------------------------------------------------------------


def summary_to_dict(summary: EvalSummary) -> dict[str, Any]:
    """把 EvalSummary 序列化为 job.result 用的 dict。

    per_case 按 case 嵌套：``[{case_id, groups: {gname: {per_metric, score}}}]``；
    aggregate 即 ``{gname: {...}}``；overall 为跨组综合（``_overall`` 总均值）。
    """
    by_case: dict[str, dict[str, Any]] = {}
    for cs in summary.per_case:
        slot = by_case.setdefault(cs.case_id, {"case_id": cs.case_id, "groups": {}})
        slot["groups"][cs.group] = {
            "per_metric": cs.per_metric,
            "score": cs.score,
        }
    return {
        "per_case": list(by_case.values()),
        "aggregate": {gname: dict(scores) for gname, scores in summary.aggregate.items()},
        "overall": dict(summary.overall),
        "extraction_summary": dict(summary.extraction_summary),
    }


async def run_offline_eval(
    job: Job,
    request: OfflineEvalRequest,
) -> EvalSummary | None:
    """三阶段执行：ingest → scoring → aggregate，推送进度并写入 job.result。

    成功返回 ``EvalSummary``；普通异常吞入 ``job.status=FAILED`` 并返回
    ``None``（避免后台 task 产生 unretrieved-exception 警告）；取消重抛。
    """
    job.status = JobStatus.RUNNING
    try:
        scorer = OfflineMetricScorer(request.groups)
        total = len(request.raw_records) * len(request.groups)
        job.push_event("progress", {"phase": "ingest", "done": 0, "total": total})

        def _ingest_progress(done: int, total_n: int) -> None:
            job.push_event("progress", {"phase": "ingest", "done": done, "total": total_n})

        materialized = await ingest(
            request.raw_records,
            request.groups,
            request.id_field,
            on_progress=_ingest_progress,
        )

        judged_labels = await _judge(job, materialized, request.groups, request.model)

        summary = await _score_and_summarize(
            job, materialized, request.groups, scorer, judged_labels
        )

        job.result = summary_to_dict(summary)
        job.status = JobStatus.COMPLETED
        job.push_event("completed", {"status": "completed"})
        return summary
    except asyncio.CancelledError:
        print(f"[OFFLINE EVAL CANCELLED] job_id={job.job_id}", flush=True)
        raise
    except Exception as e:
        tb = traceback.format_exc()
        print(f"[OFFLINE EVAL FAILED] {type(e).__name__}: {e}\n{tb}", flush=True)
        job.status = JobStatus.FAILED
        job.error = f"{type(e).__name__}: {e}"
        job.push_event("error", {"status": "failed", "error": job.error})
        return None


async def _score_and_summarize(
    job: Job,
    materialized: list[MaterializedCase],
    groups: list[GroupConfig],
    scorer: OfflineMetricScorer,
    judged_labels: dict[int, str],
) -> EvalSummary:
    """scoring → aggregate 两阶段，推送进度，返回 EvalSummary。

    ``judged_labels`` 是 llm_judge 阶段算好的 ``{materialized_index: label}``；先把它
    写回 materialized（``replace`` 重建，frozen dataclass），scoring 循环再经
    ``judged_label`` kwarg 注入 per-case 指标。其它组忽略。
    """
    total = len(materialized)
    job.push_event("progress", {"phase": "scoring", "done": 0, "total": total})
    judge_group_names = {g.name for g in groups if g.kind == "llm_judge"}
    # 写回 judged_label：llm_judge 组缺省 "其他"（兜底诊断桶），非 llm_judge 留 ""。
    materialized = [
        replace(
            mc,
            judged_label=judged_labels.get(
                i, "其他" if mc.group in judge_group_names else ""
            ),
        )
        for i, mc in enumerate(materialized)
    ]
    case_scores: list[CaseScore] = []
    for index, mc in enumerate(materialized):
        kwargs: dict[str, Any] = {}
        if mc.group in judge_group_names:
            kwargs["judged_label"] = mc.judged_label
        per_metric = scorer.score_case(mc.group, mc.extracted, mc.gold, **kwargs)
        case_scores.append(
            CaseScore(
                case_id=mc.case_id,
                group=mc.group,
                per_metric=per_metric,
                score=OfflineMetricScorer.composite(per_metric),
            )
        )
        if (index + 1) % _PROGRESS_EVERY == 0 or index + 1 == total:
            job.push_event(
                "progress",
                {"phase": "scoring", "done": index + 1, "total": total},
            )

    job.push_event("progress", {"phase": "aggregate", "done": 0, "total": len(groups)})
    summary = scorer.summarize(case_scores, materialized)
    job.push_event("progress", {"phase": "aggregate", "done": len(groups), "total": len(groups)})
    return summary
