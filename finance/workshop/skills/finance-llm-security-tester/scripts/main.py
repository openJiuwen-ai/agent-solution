"""
大模型安全扫描测试主程序
实现多线程并发请求Qwen3m模型，并将结果保存到Excel
"""

import asyncio
import openpyxl
import logging
import os
import time
import json
from datetime import datetime
from typing import List, Tuple, Optional
import sys

from config import (
    INPUT_EXCEL_PATH,
    DATA_DIR,
    LOGS_DIR,
    INPUT_SHEET_NAME,
    PROMPT_COLUMN,
    RESULT_COLUMN,
    START_ROW,
    NUM_WORKERS,
    MODEL_REQUEST_CONFIG,
    SAVE_THRESHOLD, QWEN_MODEL_NAME, REQUEST_INTERVAL, SYSTEM_PROMPT
)
from qwen_client import QwenClient


class ColoredFormatter(logging.Formatter):
    """彩色日志格式化器"""

    # ANSI颜色代码
    COLORS = {
        'DEBUG': '\033[36m',      # 青色
        'INFO': '\033[32m',       # 绿色
        'WARNING': '\033[33m',    # 黄色
        'ERROR': '\033[31m',      # 红色
        'CRITICAL': '\033[35m',   # 紫色
        'RESET': '\033[0m',       # 重置
    }

    def format(self, record):
        # 保存原始的msg和args
        original_msg = record.msg
        original_args = record.args
        
        # 先获取格式化后的消息（这会处理args）
        try:
            message = record.getMessage()
        except TypeError:
            # 如果格式化失败，直接使用原始消息
            message = str(record.msg)
        
        # 获取日志级别对应的颜色
        color = self.COLORS.get(record.levelname, self.COLORS['RESET'])

        # 为成功响应添加特殊标记和颜色
        if '收到响应' in message or '成功' in message:
            color = '\033[92m'  # 亮绿色
        elif '失败' in message or '异常' in message or '错误' in message:
            color = '\033[91m'  # 亮红色

        # 格式化消息
        record.levelname = f"{color}{record.levelname}{self.COLORS['RESET']}"
        record.name = f"{color}{record.name}{self.COLORS['RESET']}"
        record.msg = f"{color}{message}{self.COLORS['RESET']}"
        record.args = ()  # 清空args，避免父类再次尝试格式化

        # 调用父类的format方法
        formatted = super().format(record)

        # 恢复原始值（避免影响其他handler）
        record.msg = original_msg
        record.args = original_args

        return formatted


class SafeFormatter(logging.Formatter):
    """安全的日志格式化器（用于文件日志，处理带参数的日志消息）"""
    
    def format(self, record):
        # 保存原始的msg和args
        original_msg = record.msg
        original_args = record.args
        
        # 先获取格式化后的消息（这会处理args）
        try:
            message = record.getMessage()
        except TypeError:
            # 如果格式化失败，直接使用原始消息
            message = str(record.msg)
        
        # 设置格式化后的消息
        record.msg = message
        record.args = ()
        
        # 调用父类的format方法
        formatted = super().format(record)
        
        # 恢复原始值（避免影响其他handler）
        record.msg = original_msg
        record.args = original_args
        
        return formatted


class RequestQueue:
    """请求队列 - 使用队列管理待处理的提示词"""
    
    def __init__(self, prompts: List[Tuple[int, str]]):
        """
        初始化请求队列
        
        Args:
            prompts: 提示词列表，格式为[(行号, 提示词), ...]
        """
        self.queue = asyncio.Queue()
        self.total_count = len(prompts)
        self.completed_count = 0
        self.pass_count = 0
        self.reject_count = 0
        self.other_count = 0
        self._lock = asyncio.Lock()  # 用于线程安全的锁
        
        # 将所有提示词放入队列
        for row, prompt in prompts:
            self.queue.put_nowait((row, prompt))
    
    async def update_stats(self, result: Optional[str]):
        """
        更新统计信息（线程安全）
        
        Args:
            result: 响应结果
        """
        async with self._lock:
            if result:
                try:
                    result_json = json.loads(result)
                    status = result_json.get("status", "")
                    if status == "pass":
                        self.pass_count += 1
                    elif status == "reject":
                        self.reject_count += 1
                    else:
                        self.other_count += 1
                except (json.JSONDecodeError, TypeError):
                    self.other_count += 1
            else:
                self.other_count += 1
    
    def get_stats_summary(self) -> str:
        """
        获取统计摘要
        
        Returns:
            统计摘要字符串
        """
        total = self.completed_count
        if total == 0:
            return "暂无数据"
        
        pass_ratio = (self.pass_count / total * 100) if total > 0 else 0
        reject_ratio = (self.reject_count / total * 100) if total > 0 else 0
        
        return f"进度: {total}/{self.total_count}, pass={self.pass_count}({pass_ratio:.2f}%), reject={self.reject_count}({reject_ratio:.2f}%), 其他={self.other_count}"


def setup_logging() -> logging.Logger:
    """
    配置日志记录，支持文件输出和彩色控制台输出

    Returns:
        配置好的Logger对象
    """
    # 确保日志目录存在
    os.makedirs(LOGS_DIR, exist_ok=True)

    # 生成日志文件名（带模型名称、时间戳）
    name = QWEN_MODEL_NAME
    if "/" in name:
        name = name.split("/")[-1]
    log_filename = f"qwen_test_{name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
    log_path = os.path.join(LOGS_DIR, log_filename)

    # 获取根logger
    root_logger = logging.getLogger()
    root_logger.setLevel(logging.INFO)

    # 清除已有的handlers
    root_logger.handlers.clear()

    # 定义日志格式
    log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    safe_formatter = SafeFormatter(log_format)

    # 1. 文件处理器 - 立即刷新，确保日志写入
    file_handler = logging.FileHandler(log_path, encoding='utf-8', mode='a')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(safe_formatter)
    # 立即刷新缓冲区
    file_handler.flush()
    root_logger.addHandler(file_handler)

    # 2. 控制台处理器 - 彩色输出
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    colored_formatter = ColoredFormatter(log_format)
    console_handler.setFormatter(colored_formatter)
    root_logger.addHandler(console_handler)

    # 立即刷新所有handlers
    for handler in root_logger.handlers:
        handler.flush()

    # 禁用httpx的INFO日志（避免带参数的日志消息导致格式化错误）
    logging.getLogger('httpx').setLevel(logging.WARNING)
    logging.getLogger('httpx._client').setLevel(logging.WARNING)

    # 返回模块的logger
    logger = logging.getLogger(__name__)
    logger.info(f"日志系统初始化完成，日志文件: {log_path}")

    return logger


def read_prompts_from_excel(excel_path: str, sheet_name: str, 
                           column: str, start_row: int) -> List[Tuple[int, str]]:
    """
    从Excel文件中读取提示词
    
    Args:
        excel_path: Excel文件路径
        sheet_name: 工作表名称
        column: 提示词所在列
        start_row: 数据开始行
        
    Returns:
        包含(行号, 提示词)的列表
    """
    logger = logging.getLogger(__name__)
    prompts = []
    
    try:
        # 加载工作簿
        wb = openpyxl.load_workbook(excel_path)
        
        # 获取指定工作表
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # 如果工作表不存在，使用第一个工作表
            ws = wb.active
            logger.warning(f"工作表 '{sheet_name}' 不存在，使用活动工作表: {ws.title}")
        
        logger.info(f"工作表: {ws.title}, 总行数: {ws.max_row}")
        
        # 读取指定列的数据
        for row in range(start_row, ws.max_row + 1):
            cell = ws[f"{column}{row}"]
            if cell.value:
                prompts.append((row, cell.value))
        
        logger.info(f"成功读取 {len(prompts)} 条提示词")
        return prompts
        
    except Exception as e:
        logger.error(f"读取Excel文件失败: {str(e)}")
        raise


def save_results_to_excel(excel_path: str, sheet_name: str,
                         results: List[Tuple[int, str]],
                         output_dir: str,
                         output_path: str = None):
    """
    将结果保存到Excel文件（使用openpyxl框架）
    读取原始Excel的所有数据，只在I列写入大模型响应内容，其他字段保持原样

    Args:
        excel_path: 原Excel文件路径
        sheet_name: 工作表名称
        results: 结果列表，格式为[(行号, 响应内容), ...]
        output_dir: 输出目录
        output_path: 输出文件路径（如果指定则追加，否则创建新文件）

    Returns:
        输出文件路径
    """
    logger = logging.getLogger(__name__)

    try:
        # 如果指定了输出路径且文件存在，则追加模式
        if output_path and os.path.exists(output_path):
            logger.info(f"追加模式：打开已存在的文件 {output_path}")
            wb = openpyxl.load_workbook(output_path)
        else:
            # 否则创建新文件，基于原始Excel
            logger.info(f"新建模式：基于原始Excel创建新文件")
            wb = openpyxl.load_workbook(excel_path)

        # 获取指定工作表
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        logger.info(f"工作表: {ws.title}, 总行数: {ws.max_row}, 总列数: {ws.max_column}")

        # 将结果更新到工作表中（只在I列写入大模型响应）
        result_col_num = openpyxl.utils.column_index_from_string(RESULT_COLUMN)

        logger.info(f"准备写入 {len(results)} 条结果到 {RESULT_COLUMN} 列")

        for row, result in results:
            # 更新指定行的I列数据
            cell = ws.cell(row=row, column=result_col_num)
            cell.value = result
            logger.debug(f"更新行 {row} 列 {RESULT_COLUMN}: {result[:100] if result else 'None'}...")

        logger.info(f"数据更新完成，共更新 {len(results)} 行")

        # 统计结果
        pass_count = 0
        reject_count = 0
        other_count = 0

        for row, result in results:
            if result:
                # 解析JSON结果
                try:
                    result_json = json.loads(result)
                    status = result_json.get("status", "")
                    if status == "pass":
                        pass_count += 1
                    elif status == "reject":
                        reject_count += 1
                    else:
                        other_count += 1
                except (json.JSONDecodeError, TypeError):
                    other_count += 1
            else:
                other_count += 1

        total_count = len(results)
        pass_ratio = (pass_count / total_count * 100) if total_count > 0 else 0
        reject_ratio = (reject_count / total_count * 100) if total_count > 0 else 0

        logger.info(f"结果统计: 总数={total_count}, pass={pass_count}({pass_ratio:.2f}%), reject={reject_count}({reject_ratio:.2f}%), 其他={other_count}")

        # 确定输出文件路径
        if not output_path:
            # 生成输出文件名（带结束时间）
            end_time = datetime.now().strftime('%Y%m%d_%H%M%S')
            original_filename = os.path.basename(excel_path)
            model_name = QWEN_MODEL_NAME
            if "/" in model_name:
                model_name = model_name.split("/")[-1]
            output_filename = f"{os.path.splitext(original_filename)[0]}_{model_name}_{end_time}.xlsx"
            output_path = os.path.join(output_dir, output_filename)

        # 确保输出目录存在
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        # 使用openpyxl保存文件（保留所有格式和公式）
        wb.save(output_path)
        wb.close()

        logger.info(f"结果已保存到: {output_path}")
        logger.info(f"处理了 {len(results)} 条结果")
        logger.info(f"原始数据已完整保留，仅在 {RESULT_COLUMN} 列写入大模型响应")

        return output_path

    except Exception as e:
        logger.error(f"保存结果失败: {str(e)}")
        raise


async def worker_task(worker_id: int, qwen_client: QwenClient,
                     request_queue: RequestQueue, results_list: List[Tuple[int, str]],
                     save_callback):
    """
    工作协程任务 - 从队列中获取提示词并处理

    Args:
        worker_id: 工作协程ID
        qwen_client: Qwen客户端
        request_queue: 请求队列
        results_list: 结果列表（用于存储处理结果）
        save_callback: 保存回调函数
    """
    logger = logging.getLogger(__name__)
    logger.info(f"工作协程 {worker_id} 启动")

    while True:
        try:
            # 从队列中获取提示词（阻塞等待）
            row, prompt = await request_queue.queue.get()

            logger.debug(f"工作协程 {worker_id} 开始处理行 {row}")

            try:
                # 发送请求获取响应
                response = await qwen_client.send_request(prompt, max_retries=2)
                await asyncio.sleep(REQUEST_INTERVAL)

                if response is None:
                    logger.warning(f"工作协程 {worker_id} - 行 {row}: 请求失败，response为None")
                    response = "请求失败"
                else:
                    logger.info(f"工作协程 {worker_id} - 行 {row}: 收到响应，长度={len(response)}")

                # 将结果添加到结果列表
                results_list.append((row, response))
                logger.debug(f"工作协程 {worker_id} - 行 {row}: 结果已添加到列表，当前列表长度={len(results_list)}")

                # 增加完成计数并更新统计
                request_queue.completed_count += 1
                await request_queue.update_stats(response)

                # 打印实时统计
                logger.info(f"工作协程 {worker_id} 完成行 {row}，{request_queue.get_stats_summary()}")

                # 检查是否需要增量保存
                await save_callback()

            except Exception as e:
                logger.error(f"工作协程 {worker_id} - 行 {row}: 处理异常 - {str(e)}")
                results_list.append((row, f"处理异常: {str(e)}"))
                request_queue.completed_count += 1
                await request_queue.update_stats(None)

                # 即使异常也要检查保存
                await save_callback()

            finally:
                # 标记任务完成
                request_queue.queue.task_done()

        except asyncio.CancelledError:
            logger.info(f"工作协程 {worker_id} 被取消")
            break
        except Exception as e:
            logger.error(f"工作协程 {worker_id} 异常: {str(e)}")
            break


async def run_concurrent_tests(prompts: List[Tuple[int, str]],
                               qwen_client: QwenClient,
                               num_workers: int = NUM_WORKERS) -> List[Tuple[int, str]]:
    """
    并发运行测试 - 使用固定数量的工作协程顺序处理请求

    Args:
        prompts: 提示词列表
        qwen_client: Qwen客户端
        num_workers: 工作协程数量（默认4个）

    Returns:
        结果列表
    """
    logger = logging.getLogger(__name__)
    results = []
    output_path = None  # 输出文件路径
    last_save_count = 0  # 上次保存时的结果数量
    save_lock = asyncio.Lock()  # 保存锁，防止并发保存

    logger.info(f"开始处理 {len(prompts)} 个请求")
    logger.info(f"工作协程数量: {num_workers}")
    logger.info(f"每个协程顺序处理请求，每次只处理1个，完成后处理下一个")
    logger.info(f"增量保存阈值: {SAVE_THRESHOLD} 条")

    # 创建请求队列
    request_queue = RequestQueue(prompts)

    async def incremental_save():
        """增量保存回调函数"""
        nonlocal output_path, last_save_count

        async with save_lock:
            current_count = len(results)
            # 检查是否达到保存阈值
            if current_count - last_save_count >= SAVE_THRESHOLD:
                logger.info(f"达到保存阈值 ({current_count - last_save_count} >= {SAVE_THRESHOLD})，开始增量保存...")

                try:
                    # 保存当前所有结果
                    output_path = save_results_to_excel(
                        INPUT_EXCEL_PATH,
                        INPUT_SHEET_NAME,
                        results.copy(),  # 使用副本避免并发问题
                        DATA_DIR,
                        output_path
                    )
                    last_save_count = current_count
                    logger.info(f"增量保存完成，已保存 {current_count} 条结果")
                except Exception as e:
                    logger.error(f"增量保存失败: {str(e)}")

    # 创建工作协程任务列表
    worker_tasks = []
    for i in range(num_workers):
        task = asyncio.create_task(
            worker_task(i + 1, qwen_client, request_queue, results, incremental_save)
        )
        worker_tasks.append(task)

    # 等待所有任务完成
    await request_queue.queue.join()

    # 取消所有工作协程
    for task in worker_tasks:
        task.cancel()

    # 等待所有协程退出
    await asyncio.gather(*worker_tasks, return_exceptions=True)

    # 最终保存（保存所有剩余结果）
    if len(results) > last_save_count:
        logger.info(f"最终保存剩余 {len(results) - last_save_count} 条结果...")
        output_path = save_results_to_excel(
            INPUT_EXCEL_PATH,
            INPUT_SHEET_NAME,
            results,
            DATA_DIR,
            output_path
        )

    successful = sum(1 for _, r in results if r and not r.startswith("请求失败") and not r.startswith("处理异常"))
    logger.info(f"处理完成: 成功 {successful}/{len(prompts)}")

    return results


async def main():
    """主函数"""
    # 配置日志
    logger = setup_logging()
    logger.info("=" * 60)
    logger.info("大模型安全扫描测试程序启动")
    logger.info("=" * 60)
    
    try:
        # 创建Qwen客户端
        qwen_client = QwenClient()
        logger.info(f"客户端初始化完成，API地址: {qwen_client.api_url}")
        # 打印请求配置信息
        logger.info(f"模型请求配置: {MODEL_REQUEST_CONFIG}")
        logger.info(f"模型请求系统提示词: {SYSTEM_PROMPT}")
        
        # 从Excel读取提示词
        logger.info(f"开始读取Excel文件: {INPUT_EXCEL_PATH}")
        prompts = read_prompts_from_excel(
            INPUT_EXCEL_PATH,
            INPUT_SHEET_NAME,
            PROMPT_COLUMN,
            START_ROW
        )
        
        if not prompts:
            logger.warning("没有读取到任何提示词，程序退出")
            return
        
        # 并发处理所有提示词（使用NUM_WORKERS个协程）
        start_time = time.time()
        results = await run_concurrent_tests(prompts, qwen_client, num_workers=NUM_WORKERS)
        end_time = time.time()

        # 统计信息
        total_time = end_time - start_time
        logger.info(f"总耗时: {total_time:.2f} 秒")
        logger.info(f"平均每个请求: {total_time / len(prompts):.2f} 秒")
        logger.info(f"吞吐量: {len(prompts) / total_time:.2f} 请求/秒")

        logger.info("=" * 60)
        logger.info("测试完成！")
        logger.info("=" * 60)
        
    except Exception as e:
        logger.error(f"程序执行失败: {str(e)}")
        raise


if __name__ == "__main__":
    # 运行主程序
    asyncio.run(main())
